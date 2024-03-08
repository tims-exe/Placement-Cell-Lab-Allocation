from email.message import EmailMessage
import ssl
import smtplib
import openpyxl
from openpyxl import Workbook, load_workbook

print('Executing....')

#Available Labs
labs = {
    'Power Systems Lab' : 40,
    'Mech CAD Lab' : 60,
    'Van Rossum Lab': 35,
    'Hinton Lab' : 30,
    'John McCarthy Lab' : 35,
    'Foss Lab': 40
}

#excel file form google sheets
excel_file = 'test_labs.xlsx'

#set the coloumn letter of each from the excel file
excel_lab = 'D'
excel_name = 'C'
excel_email = 'B'

#loading excel sheet
book = load_workbook(excel_file)
sheet = book.active

#variables
#change email and password
count = 0
lab_index = 2
sender = 'dummy@gmail.com'
password = 'dummy passord'


class User:
    def __init__(self, name, email):
        self.name = name
        self.email = email
    def set_lab(self, lab_name):
        self.lab = lab_name
        print(self.name, 'is allocated to ',lab_name)
    
    def send_email(self):
        subject = "Placement Cell"

        body = f"""
        you are requested to go to {self.lab} for your exam
        """

        em = EmailMessage()

        em['From'] = sender
        em['To'] = self.email
        em['Subject'] = subject
        em.set_content(body)

        context = ssl.create_default_context()

        try :
            with smtplib.SMTP_SSL('smtp.gmail.com', 465, context = context) as smtp:
                smtp.login(sender, password)
                smtp.sendmail(sender, self.email, em.as_string())
            print('\nEmail Sent!')
        except:
            print('\nEmail Failed to Send')
        
sheet[excel_lab+str(1)].value = 'Labs'

for i in range(2, sheet.max_row+1):
    person = User(sheet[excel_name+str(i)].value, sheet[excel_email+str(i)].value)
    if (len(labs) == 0):
        print('labs are full, ',sheet['C'+str(i)].value,'is not allocated')
    else :
        for j in labs:
            if (labs[j] > 0):
                person.set_lab(j)
                sheet[excel_lab+str(lab_index)].value = j
                lab_index += 1
                labs[j] -= 1
                count += 1
                break
            else :
                print(j ,'is fully allocated with count = ', count, '\n')
                del labs[j]
                count = 0
                break
book.save('lab_assigned.xlsx')
